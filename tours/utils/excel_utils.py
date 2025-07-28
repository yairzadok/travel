import openpyxl
from django.http import HttpResponse
from tours.models import TravelerRegistration, Tour
from django.utils.text import slugify

def export_registrations_excel(tour_id=None):
    """
    יוצר קובץ Excel עם רשימת נרשמים.
    אם ניתן tour_id, מייצא רק את אותו סיור – עם שם הטיול והמדריך בכותרת.
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if tour_id:
        tour = Tour.objects.get(pk=tour_id)
        filename = f"{slugify(tour.tour_title)}_{tour.id}_registrations.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    else:
        response['Content-Disposition'] = 'attachment; filename="all_registrations.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "נרשמים"

    row_index = 1

    # כותרת ראשית: שם הטיול והמדריך (אם יש)
    if tour_id:
        guide_name = ""
        if tour.tour_guide_name:
            guide_name = f"{tour.tour_guide_name.tour_guide_first_name} {tour.tour_guide_name.tour_guide_last_name}"

        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
        sheet.cell(row=row_index, column=1).value = f"טיול: {tour.tour_title} | מדריך: {guide_name}"
        row_index += 2  # דילוג שורה לאחר הכותרת

    # כותרות עמודות
    headers = ['שם פרטי', 'שם משפחה', 'טלפון', 'אימייל', 'נוכחות']
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=row_index, column=col_index).value = header
    row_index += 1

    queryset = TravelerRegistration.objects.all()
    if tour_id:
        queryset = queryset.filter(tour_id=tour_id)

    for r in queryset:
        row = [
            r.traveler_first_name,
            r.traveler_last_name,
            r.traveler_phone,
            r.traveler_email,
            'הגיע' if r.traveler_attendance else 'לא הגיע',
        ]
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index).value = value
        row_index += 1

    workbook.save(response)
    return response
